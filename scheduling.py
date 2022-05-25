import numpy as np, pandas as pd
from pulp import *
from ortoolpy import addvars, addbinvars
import openpyxl

#従業員氏名、技能スコア、当月稼働日数取得
list_employee = pd.read_excel('ShiftParams.xlsx', sheet_name='技能スコア', header=0, index_col=None)
print(list_employee)

#従業員数
num_employee = len(list_employee)
#print(num_employee)

#平均技能スコア
avg_skill = list_employee['技能スコア'].mean()
#print(avg_skill)

#希望休日、必須出勤日取得
preferred = pd.read_excel('ShiftParams.xlsx', sheet_name='work', header=None, index_col=None)
#print(preferred)

#カラム名取得
tmp_colname = preferred.loc[2]
#print(preferred.loc[2])

#行削除
preferred.drop(index=[0,1,2], inplace=True)
#print(preferred)

#行番号振り直し
preferred.reset_index(drop=True, inplace=True)

#カラム名設定
preferred.rename(columns = tmp_colname, inplace=True)
print(preferred)

#必要人数 Nanを0に変換
preferred.必要人数 = preferred.必要人数.replace(np.nan, 0)

#必要人数のみ抽出
req_numpeople = pd.DataFrame(preferred.必要人数)
#print(req_numpeople)

#従業員希望列のみ抽出
emp_preferred = preferred.iloc[:,3:]
#print(emp_preferred)


list_dateno = list(range(len(emp_preferred)))
list_empno = list(range(num_employee))

model = LpProblem() #数理モデル
shift = pd.DataFrame(addbinvars(len(list_dateno), len(list_empno))) #0-1変数リスト

# 必要人数一致制約、人数案分
cost_empdif = 500
con_empdif = addvars(len(emp_preferred)) #非負変数リスト
for _,r in req_numpeople[3:].iterrows(): #0-2行目は前月のため非処理
    #print(_)
    #print(r.必要人数)
    if(r.必要人数 == 0):
        model += lpSum(shift.loc[_]) + con_empdif[_] >= (num_employee / 2)
        model += lpSum(shift.loc[_]) - con_empdif[_] <= (num_employee - 1)
    else:
        model += lpSum(shift.loc[_]) == r.必要人数

# 出勤日数の一致制約
for i in list_empno:
    model += lpSum(shift.iloc[3:, i]) == list_employee.at[i, '稼働日数'] #データフレーム従業員行、稼働日数列

# 休日希望、必須出勤の一致制約
for i in list_empno:
    for j in list_dateno:
        if emp_preferred.iloc[j,i] == 1:
            model += shift.iloc[j,i] == 1 #必須出勤
        elif emp_preferred.iloc[j,i] == 0:
            model += shift.iloc[j,i] == 0 #希望休日

# 管理者不足に対してペナルティーを求める。(少なくともAAA、BBBの何れかが出勤することが望ましい)
cost_manage = 100
con_manage = addbinvars(len(emp_preferred))
for _,i in shift[3:].iterrows():
    model += lpSum(i[0:2]) + con_manage[_] >= 1

# 技能不足に対してペナルティーを求める。
cost_skill = 50
con_skill = addbinvars(len(emp_preferred)) #0-1変数リスト
for _,r in shift[3:].iterrows():
    model += lpSum(lpDot(r,list_employee['技能スコア'])) + con_skill[_] >= lpSum(r) * avg_skill

# 4連勤に対してペナルティーを求める。
cost_work3  = 1000
con_work3 = np.array(addbinvars(len(emp_preferred)-3, num_employee)) #0-1変数リスト
for i in list_empno:
    for n,p in enumerate((shift.values[:-3,i] + shift.values[1:-2,i] + shift.values[2:-1,i] + shift.values[3:,i]).flat):
        model += p - con_work3[n][i] <= 3

# 3連休に対してペナルティーを求める。
cost_rest2 = 1000
con_rest2 = np.array(addbinvars(len(emp_preferred)-2, num_employee)) #0-1変数リスト
for i in list_empno:
    for n,p in enumerate((shift.values[:-2,i] + shift.values[1:-1,i] + shift.values[2:,i]).flat):
        model += p + con_rest2[n][i] >= 1

# 目的関数
model += (
    + cost_work3 * lpSum(con_work3)
    + cost_rest2 * lpSum(con_rest2)
    + cost_manage * lpSum(con_manage)
    + cost_skill * lpSum(con_skill)
    + cost_empdif * lpSum(con_empdif)
)

model.solve()
result = np.vectorize(value)(shift).astype(int) #0に近い程、条件を満たした最適な組み合わせ。
#print(result)
print('目的関数', value(model.objective))


# 整合チェック、結果出力データ生成

# 希望休日、希望出勤チェック
chk_diff0 = np.zeros((len(emp_preferred), num_employee))
chk_diff1 = np.zeros((len(emp_preferred), num_employee))
DF_chk_diff0 = pd.DataFrame(data=chk_diff0, index=None, columns=None)
DF_chk_diff1 = pd.DataFrame(data=chk_diff1, index=None, columns=None)
DF_chk_diff0.replace(0, np.nan, inplace=True)
DF_chk_diff1.replace(0, np.nan, inplace=True)

for i in list_dateno:
    for j in list_empno:
        # 不正値の場合
        if not (value(shift.iloc[i,j]) == 1.0) and not (value(shift.iloc[i,j]) == 0.0):
            print(value(shift.iloc[i,j]))
        # 出勤割当かつ希望休日の場合
        elif (result[i][j] == 1) and (emp_preferred.iloc[i,j] == 0):
            print(str(i) +"日目")
            print(preferred.columns[j])
            DF_chk_diff0.iloc[i,j] = 1
        # 休暇割当かつ希望出勤の場合
        elif (result[i][j] == 0) and (emp_preferred.iloc[i,j] == 1):
            print(str(i) +"日目")
            print(preferred.columns[j])
            DF_chk_diff1.iloc[i,j] = 1

#print(DF_chk_diff0)
#print(DF_chk_diff1)

# 4連勤チェック、3連休チェック
chk_work3 = np.zeros((len(emp_preferred), num_employee))
chk_rest2 = np.zeros((len(emp_preferred), num_employee))
DF_chk_work3 = pd.DataFrame(data=chk_work3, index=None, columns=None)
DF_chk_rest2 = pd.DataFrame(data=chk_rest2, index=None, columns=None)
DF_chk_work3.replace(0, np.nan, inplace=True)
DF_chk_rest2.replace(0, np.nan, inplace=True)

for i in list_dateno[3:]:
    for j in list_empno:
        if((result[i-3][j] + result[i-2][j] + result[i-1][j] + result[i][j]) > 3):
            DF_chk_work3.iloc[i,j] = result[i-3][j] + result[i-2][j] + result[i-1][j] + result[i][j]

for i in list_dateno[2:]:
    for j in list_empno:
        if((result[i-2][j] + result[i-1][j] + result[i][j]) < 1):
            DF_chk_rest2.iloc[i,j] = (3 - (result[i-2][j] + result[i-1][j] + result[i][j]))

#print(DF_chk_work3)
#print(DF_chk_rest2)

# 管理者不在チェック
chk_manage = np.zeros((len(emp_preferred), num_employee))
DF_chk_manage = pd.DataFrame(data=chk_manage, index=None, columns=None)
DF_chk_manage.replace(0, np.nan, inplace=True)

for i in list_dateno:
    if (result[i][0] == 0) and (result[i][1] == 0):
        print("管理者不在")
        print(str(i) +"日目")
        DF_chk_manage.iloc[i,0] = 1
        DF_chk_manage.iloc[i,1] = 1

#print(DF_chk_manage)

# 結果出力
path = 'ShiftParams.xlsx'
wb = openpyxl.load_workbook(path)
if '結果' in wb.sheetnames:
    wb.remove(wb['結果'])

if '休日不一致' in wb.sheetnames:
    wb.remove(wb['休日不一致'])

if '出勤不一致' in wb.sheetnames:
    wb.remove(wb['出勤不一致'])

if '連勤不一致' in wb.sheetnames:
    wb.remove(wb['連勤不一致'])

if '連休不一致' in wb.sheetnames:
    wb.remove(wb['連休不一致'])

if '管理者不在' in wb.sheetnames:
    wb.remove(wb['管理者不在'])

wb.save(path)

with pd.ExcelWriter(path, mode='a') as writer:
    df0 = pd.DataFrame(result)
    df0.columns = emp_preferred.columns
    df0.insert(0, '曜日', preferred['曜日'])
    df0['当日人数'] = 0
    df0['必要人数'] = req_numpeople
    df0['必要人数差'] = 0
    df0['目的関数'] = np.nan
    df0.at[3, '目的関数'] = value(model.objective)

    for index, row in df0.iterrows():
        #print(index)
        #print(row)
        df0.at[index,'当日人数'] = sum(row[1:-3])
        if(row.必要人数 != 0.0):
            df0.at[index,'必要人数差'] = df0.at[index,'必要人数'] - df0.at[index,'当日人数']
        else:
            df0.at[index,'必要人数差'] = 0

    df0.index = preferred['日付']

    df0[3:].to_excel(writer, sheet_name='結果')

    df1 = pd.DataFrame(DF_chk_diff0)
    df1.columns = emp_preferred.columns
    df1.insert(0, '曜日', preferred['曜日'])

    df1.index = preferred['日付']
    #print(df1)

    df1[3:].to_excel(writer, sheet_name='休日不一致')

    df2 = pd.DataFrame(DF_chk_diff1)
    df2.columns = emp_preferred.columns
    df2.insert(0, '曜日', preferred['曜日'])

    df2.index = preferred['日付']
    #print(df2)

    df2[3:].to_excel(writer, sheet_name='出勤不一致')

    df3 = pd.DataFrame(DF_chk_work3)
    df3.columns = emp_preferred.columns
    df3.insert(0, '曜日', preferred['曜日'])

    df3.index = preferred['日付']
    #print(df3)

    df3[3:].to_excel(writer, sheet_name='連勤不一致')

    df4 = pd.DataFrame(DF_chk_rest2)
    df4.columns = emp_preferred.columns
    df4.insert(0, '曜日', preferred['曜日'])

    df4.index = preferred['日付']
    #print(df4)

    df4[3:].to_excel(writer, sheet_name='連休不一致')

    df5 = pd.DataFrame(DF_chk_manage)
    df5.columns = emp_preferred.columns
    df5.insert(0, '曜日', preferred['曜日'])

    df5.index = preferred['日付']
    #print(df5)

    df5[3:].to_excel(writer, sheet_name='管理者不在')
